import os
import logging
from datetime import datetime, date
from pathlib import Path

from openpyxl import load_workbook

logger = logging.getLogger(__name__)

WORKBOOK_PATH = os.getenv("EXCEL_PATH", "Career Tracker.xlsx")

TO_APPLY_SHEET = "To Apply"
APPLICATION_SHEET = "Application Tracker"

TO_APPLY_HEADERS = [
    "Gmail Message ID", "Company", "Position", "Location",
    "Apply Link", "Source", "Industry", "Date Found", "Status",
]

# Column indices in To Apply sheet (0-based)
_COL_MSG_ID   = 0
_COL_COMPANY  = 1
_COL_POSITION = 2
_COL_LOCATION = 3
_COL_LINK     = 4
_COL_SOURCE   = 5
_COL_INDUSTRY = 6
_COL_DATE     = 7
_COL_STATUS   = 8


def _load_wb():
    path = Path(WORKBOOK_PATH)
    if not path.exists():
        raise FileNotFoundError(
            f"'{WORKBOOK_PATH}' not found. Place your Excel file in the project folder."
        )
    return load_workbook(WORKBOOK_PATH)


def _save_wb(wb):
    try:
        wb.save(WORKBOOK_PATH)
    except PermissionError:
        raise RuntimeError(
            f"Cannot save — close '{WORKBOOK_PATH}' in Excel and try again."
        )


def ensure_to_apply_sheet():
    """Create the 'To Apply' sheet with headers if it doesn't exist yet."""
    wb = _load_wb()
    if TO_APPLY_SHEET not in wb.sheetnames:
        ws = wb.create_sheet(TO_APPLY_SHEET)
        ws.append(TO_APPLY_HEADERS)
        _save_wb(wb)
        logger.info("Created 'To Apply' sheet.")
    wb.close()


def get_existing_job_keys() -> set:
    """
    Return a set of (company_lower, position_lower) from both
    'To Apply' and 'Application Tracker' for fast duplicate checks.
    """
    keys = set()
    try:
        wb = _load_wb()

        if TO_APPLY_SHEET in wb.sheetnames:
            ws = wb[TO_APPLY_SHEET]
            for row in ws.iter_rows(min_row=2, values_only=True):
                c = str(row[_COL_COMPANY] or "").lower().strip()
                p = str(row[_COL_POSITION] or "").lower().strip()
                if c:
                    keys.add((c, p))

        ws_app = wb[APPLICATION_SHEET]
        headers = [cell.value for cell in ws_app[1]]
        try:
            ci = headers.index("Company")
            pi = headers.index("Position")
        except ValueError:
            wb.close()
            return keys

        for row in ws_app.iter_rows(min_row=2, values_only=True):
            c = str(row[ci] or "").lower().strip()
            p = str(row[pi] or "").lower().strip()
            if c:
                keys.add((c, p))

        wb.close()
    except Exception as e:
        logger.error(f"get_existing_job_keys error: {e}")
    return keys


def add_to_apply_batch(jobs: list) -> int:
    """
    Add multiple parsed jobs to the 'To Apply' sheet in a single save.
    Returns the count of jobs written.
    """
    wb = _load_wb()
    if TO_APPLY_SHEET not in wb.sheetnames:
        ws = wb.create_sheet(TO_APPLY_SHEET)
        ws.append(TO_APPLY_HEADERS)
    else:
        ws = wb[TO_APPLY_SHEET]

    today = date.today().isoformat()
    for job in jobs:
        ws.append([
            job.get("message_id", ""),
            job.get("company", ""),
            job.get("position", ""),
            job.get("location", ""),
            job.get("apply_link", ""),
            job.get("source", ""),
            job.get("industry", ""),
            job.get("date_found", today),
            "To Apply",
        ])

    _save_wb(wb)
    wb.close()
    return len(jobs)


def get_to_apply_jobs() -> list:
    """
    Return all jobs from 'To Apply' that have not been marked as Applied.
    Each job is a dict keyed by the sheet headers.
    """
    try:
        wb = _load_wb()
        if TO_APPLY_SHEET not in wb.sheetnames:
            wb.close()
            return []

        ws = wb[TO_APPLY_SHEET]
        headers = [cell.value for cell in ws[1]]
        jobs = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            job = dict(zip(headers, row))
            if str(job.get("Status", "")).lower() == "applied":
                continue
            jobs.append(job)

        wb.close()
        return jobs
    except Exception as e:
        logger.error(f"get_to_apply_jobs error: {e}")
        raise RuntimeError(f"Could not read 'To Apply' sheet: {e}")


def get_applications() -> list:
    """
    Return all rows from 'Application Tracker', newest first.
    Datetime values are converted to ISO date strings.
    """
    try:
        wb = _load_wb()
        ws = wb[APPLICATION_SHEET]

        headers = [cell.value for cell in ws[1]]
        # Name any blank header columns so the dict doesn't lose data
        headers = [h if h else f"_col{i}" for i, h in enumerate(headers)]

        apps = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            app = {}
            for key, val in zip(headers, row):
                if isinstance(val, datetime):
                    val = val.strftime("%Y-%m-%d")
                app[key] = val
            apps.append(app)

        wb.close()

        apps.sort(
            key=lambda r: r.get("Date Applied") or "0000-00-00",
            reverse=True,
        )
        return apps
    except Exception as e:
        logger.error(f"get_applications error: {e}")
        raise RuntimeError(f"Could not read 'Application Tracker': {e}")


def mark_applied(message_id: str, form_data: dict) -> None:
    """
    1. Find the job in 'To Apply' by message_id and mark it Applied.
    2. Append a new row to 'Application Tracker'.
    Raises RuntimeError on failure (e.g. file locked, job not found).
    """
    wb = _load_wb()

    # ── Step 1: locate & update in To Apply ──
    job = None
    if TO_APPLY_SHEET in wb.sheetnames:
        ws_ta = wb[TO_APPLY_SHEET]
        for row in ws_ta.iter_rows(min_row=2):
            if str(row[_COL_MSG_ID].value) == str(message_id):
                job = {
                    "company":  row[_COL_COMPANY].value,
                    "position": row[_COL_POSITION].value,
                    "location": row[_COL_LOCATION].value,
                    "industry": row[_COL_INDUSTRY].value,
                }
                row[_COL_STATUS].value = "Applied"
                break

    if job is None:
        wb.close()
        raise RuntimeError(f"Job '{message_id}' not found in 'To Apply' sheet.")

    # ── Step 2: append to Application Tracker ──
    ws_app = wb[APPLICATION_SHEET]

    # Add "Follow-up Date" header to column 16 if it's still blank
    if ws_app.cell(row=1, column=16).value is None:
        ws_app.cell(row=1, column=16).value = "Follow-up Date"

    # Application Tracker column order (matches your existing headers):
    # Status | Company | Position | Date Applied | Role | Cover Letter |
    # Résumé upload? | Résumé Form? | Connections? | Industry | Location |
    # Date Posted | Salary Range | Latest word | Notes | Follow-up Date
    ws_app.append([
        form_data.get("status", "App Submitted"),
        job["company"],
        job["position"],
        form_data.get("date_applied", date.today().isoformat()),
        form_data.get("role", ""),
        form_data.get("cover_letter", "No"),
        form_data.get("resume_upload", "Yes"),
        form_data.get("resume_form", "No"),
        form_data.get("connections", "NIL"),
        job["industry"] or "",
        job["location"] or "",
        "",                                       # Date Posted — left blank
        "",                                       # Salary Range — left blank
        "",                                       # Latest word — user fills later
        form_data.get("notes", ""),
        form_data.get("followup_date", ""),
    ])

    _save_wb(wb)
    wb.close()
